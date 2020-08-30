# -*- coding: utf-8 -*-
"""
Created on Thu Jan 31 09:18:57 2019

@author: mlg1kor
"""

import openpyxl
file = r'D:\documents\SecurityAccess.xlsx'
wb=openpyxl.load_workbook(file)
for sheet in wb.worksheets:
  for i in range(2,sheet.max_row+1):
      for j in range(4,7):
          if sheet.cell(row=i,column=j).value!=None:
            sheet.cell(row=i,column=j).value = '<p>'+sheet.cell(row=i,column=j).value.replace('\n','</p>\n<p>')+'</p>'
wb.save(file)

