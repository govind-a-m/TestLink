# -*- coding: utf-8 -*-
"""
Created on Tue Aug  6 14:52:43 2019

@author: MLG1KOR
"""

import os
import openpyxl
import re

test_spec_folder = r'C:\Users\mlg1kor\testspec_317'
rep = {
       "&nbsp;":" ",
       "<br />":"",
       "&quot;":'"',
       "&apos;":"'",
       "&lt;":"<",
       "&amp;":"&",
       "&gt;":">",
    }
for root,_,files in os.walk(test_spec_folder):
  for file in files:
    if file.endswith('.xlsx'):
      wb=openpyxl.load_workbook(os.path.join(root,file))
      sheet = wb.active
      max_row=sheet.max_row
      for i in range(2,6):
        for j in range(1,max_row+1):
          if sheet.cell(row=j,column=i).value!=None:
            sheet.cell(row=j,column=i).value = re.sub(r'<a.+">','',sheet.cell(row=j,column=i).value)
            sheet.cell(row=j,column=i).value = re.sub(r'</?\w+>','',sheet.cell(row=j,column=i).value)
            for key,value in rep.items():
              sheet.cell(row=j,column=i).value = sheet.cell(row=j,column=i).value.replace(key,value)
      wb.save(file)

# =============================================================================
# rep = {"<p>":"",
#        "</p>":"",
#        "</a>":"",
#        "&nbsp;":" ",
#        "<br />":"",
#        "&quot;":'"',
#        "&apos;":"'",
#        "&lt;":"<",
#        "&amp;":"&",
#        "&gt;":">",
#        "</strong>":"",
#        "<strong>":"",
#        "<ol>":"",
#        "</ol>":"",
#        "<li>":"",
#        "</li>":"",
#     }
# =============================================================================
