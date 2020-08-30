# -*- coding: utf-8 -*-
"""
Created on Mon Jul 29 13:01:09 2019

@author: MLG1KOR
"""

from   selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from recurse_tl_v3 import gm_get_nodes
import json
import asil_test_class as platform
import openpyxl

node_id = '19272'
node_name = 'HW-SW ADC'
with open('tl_login_info.json','r+') as fjson:
    lg_info = json.load(fjson)
browser = webdriver.Firefox()
wait = WebDriverWait(browser, 50)
platform.browser = browser
platform.wait = wait
platform.login(lg_info)
if 'node_list' not in globals():
  nodes = gm_get_nodes(lg_info['username'],lg_info['password'],node_id,node_name)
  node_list = [value for key,value in nodes.items()]

wb=openpyxl.load_workbook(r'C:\Users\mlg1kor\Documents\My Received Files\IohWAb_ADC_mapping.xlsx')
sheet = wb.active
for nd in node_list:
  cur_tc = platform.testcase(nd,1,'QM',6)
  cur_tc.set_tcvid()
  cur_tc.create_new_version()
  cur_tc.edit()
  cur_tc.set_status()
  cur_tc.set_asil()
  cur_tc.set_tcls()
  cur_tc.add_new_req('ESW3.1_2108, ESW3.1_2834, ESW3.1_2816')
  platform.save_data()
# =============================================================================
# for i in  range(1,sheet.max_row+1):
#   tid = sheet.cell(row=i,column=1).value.strip()
#   req = sheet.cell(row=i,column=2).value.strip()
#   for j in range(len(node_list)):
#     if node_list[j].internal_id==tid:
#       break
#   nd = node_list.pop(j)
#   cur_tc = platform.testcase(nd,1,'QM',6)
#   cur_tc.set_tcvid()
#   cur_tc.create_new_version()
#   cur_tc.edit()
#   cur_tc.add_new_req(req)
#   platform.save_data()
# =============================================================================
