# -*- coding: utf-8 -*-
"""
Created on Fri Sep  6 13:32:28 2019

@author: MLG1KOR
"""

import openpyxl
import os
import time
import pyautogui
import selenium.webdriver.support.expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from collections import defaultdict


class testsuite:
  db = {}
  def __init__(self,name,node_id,tcs):
    self.name = name
    self.node_id = node_id
    self.TestCases = tcs
    self.__class__.db.update({self.node_id:self})
    
  def add_testcase(self,branch):
    TestCase = testcase(branch['name'],branch['id'],branch['tcversion_id'],branch['text'],branch['external_id'])
    self.TestCases.append(TestCase)
  
  def report_urls(self):
    li = []
    for testcase in self.TestCases:
      li.append(f'http://10.58.199.163:8025/testlink-1.9.16/lib/execute/execPrint.php?id={testcase.report_id}')
    return li
  
  def getFilenames(self,ext):
    li = []
    for testcase in self.TestCases:
      li.append(f"{testcase.report_id}.{ext}")
    return li
  
  def nameIDdict(self):
    d = {}
    for testcase in self.TestCases:
      d.update({testcase.name:testcase.node_id})
    return d
  
  def getSuiteFileName(self):
    return 'D:\\documents\\testreports\\'+self.name+'_'+str(self.node_id)+'.pdf'

  def __repr__(self):
    return f'name:{self.name},id:{self.node_id},{self.TestCases}'
  
  def CollectSteps(self,stepDB):
    for testcase in self.TestCases:
      if testcase.externalID in stepDB:
        testcase.steps = stepDB[testcase.externalID]
      else:
        print(f'did not find {testcase.name}')
        
    
class testcase:
  browser = None
  wait = None
  def __init__(self,name,node_id,tcversion_id,result,external_id):
    self.name = name
    self.node_id = node_id
    self.tcversion_id = tcversion_id
    self.result = result
    self.externalID = external_id
    self.report_id = None
    self.steps = []
    self.status = True
    
  def VisitTestCase(self,build):
    self.browser.get(f'http://10.58.199.163:8025/testlink-1.9.16/lib/execute/execSetResults.php?version_id={self.tcversion_id}&level=testcase&id={self.node_id}&setting_build={build}')
    self.wait.until(EC.visibility_of_element_located((By.XPATH,'/html/body/div/form/div[11]/div[1]/table/tbody')))
        
  def set_report_id(self):
    report_id = self.browser.find_element_by_xpath('/html/body/div/form/div[9]/table/tbody/tr[2]/td[4]').get_attribute('title')
    self.report_id = report_id[4:-1]
    print(self.report_id)

  def downloadReport(self):
    name = str(self.report_id)+'.html'
    link = f'http://10.58.199.163:8025/testlink-1.9.16/lib/execute/execPrint.php?id={self.report_id}'
    if name in os.listdir():
      os.remove(name)
    self.browser.get(link)
    time.sleep(1)
    pyautogui.hotkey('ctrl', 's')
    time.sleep(1)
    pyautogui.typewrite(name)
    pyautogui.hotkey('enter')
    time.sleep(1)

  def savedata(self):
    if self.status:
      xpath = f'//*[@id="fastExecp_{self.tcversion_id}"]'
    else:
      xpath = f'//*[@id="fastExecf_{self.tcversion_id}"]'
    self.browser.find_element_by_xpath(xpath).click()
    self.browser.find_element_by_xpath(xpath).click()
    self.browser.find_element_by_xpath('//*[@id="execute_cases"]').click()
    self.wait.until(EC.element_to_be_clickable((By.XPATH,'//*[@id="execute_cases"]')))

    
  def __repr__(self):
    return f'[name:{self.name},node_id:{self.node_id}]'

class step:
  attatchments_directory = 'D:\\documents\\current' 
  result_filename = r'D:\documents\testspecs\eScooter_bridgeDrv_Monitor_FS_7.xlsx'
  ID_col = 1
  notes_col = 10
  status_col = 11
  png_col = 12
  db = defaultdict(list)
  
  def __init__(self,eid,notes,no,stat,attatchments):
    self.externalID = eid
    self.notes = notes
    self.no = no
    self.stat = stat
    self.attatchments = attatchments
    self.stepID = None
    self.__class__.db[self.externalID].append(self)
    print(self)
  
  def attatchFiles(self):
    if self.attatchments!='':
      testcase.browser.find_element_by_id('uploadedFile_'+self.stepID).click()
      time.sleep(0.5)
      pyautogui.typewrite(self.attatchments)
      pyautogui.hotkey('enter')
      print(f'attached {self.attatchments} to step no {self.no}')
      time.sleep(0.5)
    else:
      print('no attatchments for step no {self.no}')
      
  def sendFileKey(self):
    if len(self.attatchments)>0:
      upload_el = testcase.browser.find_element_by_id('uploadedFile_'+self.stepID)
      for file in self.attatchments:
        upload_el.send_keys(file)
      
  
  @classmethod
  def CreateFromExcel(cls):
    wb = openpyxl.load_workbook(cls.result_filename)
    sheet = wb.active
    os.chdir(cls.attatchments_directory)
    for row in range(2,sheet.max_row+1):
      if sheet.cell(row=row,column=1).value!=None:
        step_no = 1
        externalID = str(sheet.cell(row=row,column=cls.ID_col).value)
        notes = sheet.cell(row=row,column=cls.notes_col).value
        status = sheet.cell(row=row,column=cls.status_col).value.strip().upper()
        if status=="PASSED":
          status = 1
        else:
          status = 2
        png = cls.FileLinks(sheet.cell(row=row,column=cls.png_col).value,externalID,step_no)
        print(f'step {externalID} created')
        step(externalID,notes,1,status,png)
      else:
        step_no+=1
        notes = sheet.cell(row=row,column=cls.notes_col).value
        status = sheet.cell(row=row,column=cls.status_col).value.strip().upper()
        if status=="PASSED":
          status = 1
        else:
          status = 2
        png = cls.FileLinks(sheet.cell(row=row,column=cls.png_col).value,externalID,step_no)
        step(externalID,notes,step_no,status,png)
 
  @classmethod
  def FileLinks(cls,value,eid,step_no):
    ret_val = []
    if value==None:
      for filename in os.listdir():
        if eid+'_'+str(step_no) in filename:
          ret_val.append(f'{cls.attatchments_directory}\\{filename}')
    else:
      if value.lower().find('none')==-1:
        value = value.split(',')
        for raw_val in value:
          if raw_val!='':
            ret_val.append(f'{cls.attatchments_directory}\\{raw_val}')
    return ret_val
  
  def __repr__(self):
    return f'       step no {self.no} status {self.stat} fileLinks:{self.attatchments}'
    
def gm_cut_the_tree(tree,name,node_id):
  for branch in tree:
    if branch['leaf'] is False:
      gm_cut_the_tree(branch['children'],branch['name'],branch['id'])
    else:
      if branch['parent_id'] not in testsuite.db:
        testsuite(name,node_id,[])
      testsuite.db[branch['parent_id']].add_testcase(branch)

# =============================================================================
# if __name__=='__main__':
#   with open(r'C:\Users\mlg1kor\Desktop\New Text Document.txt','r') as f:
#       d = json.load(f)
#   gm_cut_the_tree(d,'root','None')
# =============================================================================
if __name__=='__main__':
  step.CreateFromExcel()