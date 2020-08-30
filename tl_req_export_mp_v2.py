# -*- coding: utf-8 -*-
"""
Created on Fri Oct 18 16:46:02 2019

@author: MLG1KOR
"""

from recurse_tl_v3 import gm_get_nodes,node
from requests import session
from bs4 import BeautifulSoup
import json
import concurrent.futures as cf
import xlsxwriter
import openpyxl


with open('tl_login_info.json','r+') as fjson:
  lg_info = json.load(fjson)
payload={"tl_login":lg_info['username'],
     "tl_password":lg_info['password'],
     "CSRFName":"CSRFGuard_2113925045",
     "CSRFToken":"d661b4dfa16c21a9b594d95cce0d4e6d52ab254ec490afd357332093b6858d7c38f405c8803e263a23033fc2fa8519dbeb7b5838bfd9e05dffb5ab2c3dd1bacb",
     "action":"login.php?viewer=" }

def Cfields(node_list):
  ret_data = []
  with session() as c:
    c.post("http://10.58.199.163:8025/testlink-1.9.16/login.php",data=payload)
    for tl_node in node_list:
      tc_url='http://10.58.199.163:8025/testlink-1.9.16/lib/testcases/archiveData.php?edit=testcase&id='+tl_node.node_id
      tc=c.get(tc_url)
      try:
        soup=BeautifulSoup(tc.content)
        cfields=soup.select('#cfields_design_time')
        c_rows=cfields[0].table.find_all('tr')
        tid=c_rows[0].find_all('td')[1].text
        rid=c_rows[1].find_all('td')[1].text
        tcls  = c_rows[2].find_all('td')[1].text
        asil = c_rows[3].find_all('td')[1].text
        ret_data.append([tl_node.internal_id,tid,rid,tcls,asil])
      except:
        ret_data.append(tl_node)
  return ret_data

def segregate(node_list,packet_size):
  start = 0
  node_li = []
  while start<len(node_list):
    node_li.append(node_list[start:start+packet_size])
    start = start+packet_size
  return node_li

  
def writeToExcel(data):
  global rno
  for i in range(5):
    sheet.write(rno,i,data[i],wrap)
  rno+=1
    

def TraverseTL(node_li):
  progress = 0
  for i in range(5):
    skip_list = []
    if len(node_li)>0:
      print(f'{i+1} iteration')
      node_li = segregate(node_li,25)
      with cf.ThreadPoolExecutor(max_workers=10) as executor:
        pool = [executor.submit(Cfields,ndl) for ndl in node_li]
        for complete in cf.as_completed(pool):
          data = complete.result()
          for tc_data in data:
            if isinstance(tc_data,node):
              print(f'skipped {tc_data.node_id} {tc_data.name}')
              skip_list.append(tc_data)
            else:
              progress+=1
              print(progress,tc_data)
              yield tc_data
      node_li = skip_list
    else:
      return None

if __name__ == '__main__':
  nodeid = '19591'
  nodename = 'performance_testing'
  export_filename = 'tc_req_performance.xlsx'
  revisit = False
  if revisit==False:
    rno=0
    book=xlsxwriter.Workbook(export_filename)
    sheet=book.add_worksheet()
    wrap=book.add_format({'text_wrap':True})
    nodes = gm_get_nodes(lg_info['username'],lg_info['password'],nodeid,nodename)
    node_list = [value for key,value in nodes.items() if value.tcase]
    node_li = node_list
    for data_packet in TraverseTL(node_li):
      writeToExcel(data_packet)
    book.close()
  else:
    wb=openpyxl.load_workbook(export_filename)
    sheet = wb.active
    rno = sheet.max_row+1
    visited =  sorted(int(cell.value[3:]) for cell in sheet['A'])
    nodes = gm_get_nodes(lg_info['username'],lg_info['password'],nodeid,nodename)
    node_list = [value for key,value in nodes.items() if value.tcase]
    complete_set = set([int(tl_node.internal_id[3:]),tl_node] for tl_node in node_list)
    not_visited = complete_set - visited_set
    
    