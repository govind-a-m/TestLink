# -*- coding: utf-8 -*-
"""
Created on Tue Sep 24 11:58:11 2019

@author: MLG1KOR
"""

import openpyxl
import re

class Port:
  def __init__(self,no,data):
    self.no = no
    for key,value in data.items():
      data[key] = format(int(value,16),'016b')[::-1]
    self.data = data
    print(self)
    
  @staticmethod
  def ConstructResultsFromExcel(sheet):
    portdata = []
    for p in range(10):
      data = {}
      data.update({'PM':sheet.cell((p*21)+2,1).value})
      data.update({'P':sheet.cell((p*21)+9,1).value})
      data.update({'PMC':sheet.cell((p*21)+0,1).value})
      data.update({'PFCE':sheet.cell((p*21)+5,1).value})
      data.update({'PFC':sheet.cell((p*21)+4,1).value})
      data.update({'PIPC':sheet.cell((p*21)+8,1).value})
      data.update({'PBDC':sheet.cell((p*21)+12,1).value})
      data.update({'PIS':sheet.cell((p*21)+20,1).value})
      data.update({'PD':sheet.cell((p*21)+14,1).value})
      data.update({'PU':sheet.cell((p*21)+13,1).value})
      data.update({'PIBC':sheet.cell((p*21)+8,1).value}) 
      portdata.append(Port(p,data.copy()))
    return portdata
  
  @staticmethod
  def importdata(txt_file):
    portdata = []
    with open(txt_file,'r') as f:
      a = f.read()
    raw_data = re.findall(r'0x\w+\n',a)
    for p in range(10):
      data = {}
      data.update({'PM':raw_data[(p*21)+2][:-1]})
      data.update({'P':raw_data[(p*21)+9][:-1]})
      data.update({'PMC':raw_data[(p*21)+0][:-1]})
      data.update({'PFCE':raw_data[(p*21)+5][:-1]})
      data.update({'PFC':raw_data[(p*21)+4][:-1]})
      data.update({'PIPC':raw_data[(p*21)+8][:-1]})
      data.update({'PBDC':raw_data[(p*21)+12][:-1]})
      data.update({'PIS':raw_data[(p*21)+20][:-1]})
      data.update({'PD':raw_data[(p*21)+14][:-1]})
      data.update({'PU':raw_data[(p*21)+13][:-1]})
      data.update({'PIBC':raw_data[(p*21)+8][:-1]}) 
      portdata.append(Port(p,data.copy()))
    return portdata
  
  def __repr__(self):
    return f'{self.no} {self.data}'
  
class PortTestCases:
  exp_col = 6
  testcaseID_col = 11
  exec_notes_col = 12
  status_col = 13
  li = []
  def __init__(self,tid,portno,bittno,data):
    self.tid = tid
    self.portno = portno
    self.bittno = bittno
    self.status = True
    self.data = data
  
  def ValidateResults(self,port):
    text = ''
    for key,value in self.data.items():
      if value is not None:
        if port.data[key][self.bittno]!=value:
          self.status = False
        text = text+f'{key}{self.portno}.{self.bittno} = {port.data[key][self.bittno]}\n'
    return text
  
  @classmethod
  def WriteResultsToExcel(cls,sheet,ports):
    for rno,tc in enumerate(PortTestCases.li,start=2):
      result = tc.ValidateResults(ports[tc.portno])
      sheet.cell(row=rno,column=cls.exec_notes_col).value = result
      sheet.cell(row=rno,column=cls.status_col).value = 'Passed' if tc.status else 'Failed'
    
  @classmethod
  def CollectExpResultsFromExcel(cls,sheet):
    for i in range(2,sheet.max_row+1):
      portno,bittno = cls.getNos(sheet.cell(row=i,column=1).value)
      results = sheet.cell(row=i,column=cls.exp_col).value
      results = re.findall(r'[P]\w+\.\w+\s*[=]\s*[01]',results)
      data = data={'PM':None,'P':None,'PMC':None,'PFCE':None,'PFC':None,'PIPC':None,'PBDC':None,'PIS':None,'PD':None,'PU':None,'PIBC':None}
      for string in results:
        key = string[:string.find('.')-1]
        val = string[-1]
        data[key] = val
      tid = sheet.cell(row=i,column=cls.testcaseID_col).value
      cls.li.append(cls(tid,portno,bittno,data.copy()))
        
  @staticmethod
  def getNos(text):
    text = re.findall(r'\d\s*\.\s*\d+',text)[0]
    return [int(x.strip()) for x in text.split('.')]
  

if __name__=="__main__":
  portdata = Port.importdata(r'C:\Users\mlg1kor\Documents\My Received Files\New Text Document.txt')
  wb=openpyxl.load_workbook(r'C:\Users\mlg1kor\Documents\Copy of HW-SW (Port)_81.xlsx')
  sheet = wb.active
  PortTestCases.CollectExpResultsFromExcel(sheet)
  PortTestCases.WriteResultsToExcel(sheet,portdata)
  wb.save(r'C:\Users\mlg1kor\Documents\Copy of HW-SW (Port)_81.xlsx')
  